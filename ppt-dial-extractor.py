# This Robot to extract the ppt dials from huge detailed report

import openpyxl as excel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font

orange = PatternFill(start_color="FFFF9900",end_color="FFFF9900",fill_type="solid")
green = PatternFill(start_color="FF00FF00",end_color="FF00FF00",fill_type="solid")

import time

test_names = r"path"
A_wb = excel.load_workbook(test_names)
A_ws = A_wb["Sheet1"]

names_container = []

for ppt in range(2,A_ws.max_row+1 ) :
    cell_column_A = A_ws.cell(ppt , 1).value
    names_container.append(cell_column_A)


detailed_path = r"path"
B_wb = excel.load_workbook(detailed_path)
# All
# All(1)
B_ws = B_wb["Sheet2"]
print(names_container)
print("Phase One Passed")


# B_ws.max_row+1
for check_column in range( 2 , B_ws.max_row+1 ) :
    # Please Differentiate between cell fill and cell value
    check_Cell = B_ws.cell(check_column, 2 )
    cell_value = str(check_Cell.value)

    try :
        if any(keyword in cell_value for keyword in names_container):
            check_Cell.fill = orange

        else:
            continue

    except Exception as e:
        print(f"Error Due to : {e}")
        continue


print("Saving...")
B_wb.save(detailed_path)